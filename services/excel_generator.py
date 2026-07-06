"""Excel报销单生成服务 - 严格按照模板格式生成"""
import io
from typing import List, Dict
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExcelGenerator:
    """报销单Excel生成器 - 匹配用户提供的模板格式"""

    # 灰色表头填充（用于酒店/其他费用的列标题行）
    _GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    def __init__(self, template_bytes: bytes):
        self.wb = load_workbook(io.BytesIO(template_bytes))
        self.ws = self.wb.active

    # ------------------------------------------------------------------
    #  公共入口
    # ------------------------------------------------------------------
    def generate(
        self,
        invoices: List[Dict],
        travel_days: Dict,
        work_description: str,
    ) -> bytes:
        """根据解析结果生成报销单"""
        # ---- 1. 分类发票 ----
        transport = []   # 火车票 + 打车 + 飞机票
        hotels = []      # 酒店
        others = []      # 快递、其他、请客餐饮

        for inv in invoices:
            t = inv.get("type", "")
            if t in ("火车票", "打车", "飞机票"):
                transport.append(inv)
            elif t == "酒店":
                hotels.append(inv)
            elif t == "餐饮":
                # 餐饮发票：仅请客的才写入报销单，非请客的作为补贴替票不写入
                if inv.get("is_entertainment", False):
                    others.append(inv)
            elif t in ("快递", "其他"):
                others.append(inv)
            # 打车行程单不单独写入，仅用于配对打车发票

        # ---- 2. 生成补贴 ----
        subsidies = self._generate_subsidies(travel_days)

        # ---- 3. 排序：火车票在前、打车在后，各自按日期排 ----
        TYPE_ORDER = {"火车票": 0, "飞机票": 1, "打车": 2}
        transport.sort(key=lambda x: (
            TYPE_ORDER.get(x.get("type", ""), 9),
            x.get("date", ""),
        ))
        hotels.sort(key=lambda x: x.get("check_in_date", x.get("date", "")))
        others.sort(key=lambda x: x.get("date", ""))

        # ---- 4. 清除模板 header 以下所有行 ----
        # 保留 Row1=大标题  Row2=表头，删除 Row3（空行）
        max_r = self.ws.max_row
        if max_r > 3:
            self.ws.delete_rows(4, max_r - 3)
        # 删除第3行（模板中的空子表头）
        self.ws.delete_rows(3, 1)

        cur = 3  # 当前写入行号（从第3行开始）

        # ---- 5-A. 城际交通：火车票 + 打车 ----
        for inv in transport:
            self._write_transport_row(cur, inv)
            cur += 1

        # 空行分隔（交通块结束）
        cur += 1

        # ---- 5-B. 补贴 ----
        # 出差天数汇总摘要（写在补贴第一行的 D 列）
        summary_text = self._build_travel_summary(travel_days)

        for i, sub in enumerate(subsidies):
            self.ws.cell(row=cur, column=5, value=sub["name"])
            self.ws.cell(row=cur, column=6, value=float(sub["amount"]))
            self.ws.cell(row=cur, column=7, value=sub.get("remark", ""))
            # D 列出差天数汇总摘要只写在第一行（与模板一致）
            if i == 0 and summary_text:
                self.ws.cell(row=cur, column=4, value=summary_text)
            cur += 1

        # 空行分隔（补贴块结束）
        cur += 1

        # ---- 5-C. 城际交通小计 ----
        transport_total = sum(float(inv.get("amount", 0)) for inv in transport)
        self.ws.cell(row=cur, column=5, value="城际交通 小计")
        self.ws.cell(row=cur, column=6, value=round(transport_total, 2))
        cur += 1

        # 空行分隔（小计结束）
        cur += 1

        # ---- 5-D. 住宿费 ----
        # 写入住宿费的列标题（灰色底）
        hotel_headers = ["入住时间段", "酒店名称", "天数", "单价", "费用", "名称", "金额(元)"]
        for col_idx, h in enumerate(hotel_headers, 1):
            cell = self.ws.cell(row=cur, column=col_idx, value=h)
            cell.fill = self._GRAY_FILL
            cell.font = Font(bold=True)
        cur += 1

        hotel_total = 0.0
        for inv in hotels:
            self._write_hotel_row(cur, inv)
            hotel_total += float(inv.get("amount", 0))
            cur += 1

        # 空行分隔（酒店明细结束）
        cur += 1

        # ---- 5-E. 住宿费小计 ----
        self.ws.cell(row=cur, column=5, value="住宿费")
        self.ws.cell(row=cur, column=6, value=round(hotel_total, 2))
        cur += 1

        # 空行分隔（住宿费小计结束）
        cur += 1

        # ---- 5-F. 其他费用 ----
        # 写入其他费用的列标题（灰色底）
        other_headers = ["日期", "其他费用名称", "", "内容", "费用名称", "金额(元)", "备注(超标原因/替票、替票原因）"]
        for col_idx, h in enumerate(other_headers, 1):
            cell = self.ws.cell(row=cur, column=col_idx, value=h)
            cell.fill = self._GRAY_FILL
            cell.font = Font(bold=True)
        cur += 1

        other_total = 0.0
        for inv in others:
            self._write_other_row(cur, inv)
            other_total += float(inv.get("amount", 0))
            cur += 1

        # 空行分隔（其他费用明细结束）
        cur += 1

        # ---- 5-G. 总合计 ----
        subsidy_total = sum(float(s["amount"]) for s in subsidies)
        grand_total = transport_total + subsidy_total + hotel_total + other_total
        self.ws.cell(row=cur, column=1, value="总合计")
        self.ws.cell(row=cur, column=6, value=round(grand_total, 2))
        cur += 1

        # ---- 5-H. 页脚 ----
        self.ws.cell(row=cur, column=1, value="已申请备用金     元")
        cur += 1
        self.ws.cell(row=cur, column=1, value='提报内容（已提报的打"√"）')
        self.ws.cell(row=cur, column=3,
                      value="□出差计划表   □行程评估表  □会议资料  □跟台总结表")
        cur += 1
        self.ws.cell(row=cur, column=1,
                      value='备注\t部门主管及负责人对提报内容打"√"项需进行指导及审核；')
        cur += 1
        self.ws.cell(row=cur, column=1,
                      value="  审批:                   会计:                                          报销人:")

        # ---- 6. 输出 ----
        out = io.BytesIO()
        self.wb.save(out)
        return out.getvalue()

    # ------------------------------------------------------------------
    #  出差天数计算
    # ------------------------------------------------------------------
    def calculate_travel_days(
        self, invoices: List[Dict], home_city: str = "杭州"
    ) -> Dict:
        """
        计算出差天数（多城市独立日期算法）

        核心思路：
        1. 每个城市独立收集“人在该城市”的日期证据
        2. 滴滴打车、火车到达、火车出发都算证据
        3. 过渡日可同时属于多个城市（如温州出发去永康，06-10同时算温州和永康）
        4. 连续日期分段，间隔≥3天视为新出差
        5. 每段天数 = 最后一天 - 第一天 + 1

        Returns:
            {地区简称: 天数}  例如 {"温州": 5, "桐庐": 3}
        """
        # ====== 第一步：收集每日城市证据 ======
        date_cities: Dict = {}  # {date: {city1, city2, ...}}
        date_no_city: Dict = {}  # {date: [inv, ...]} 无城市提取的滴滴记录

        for inv in invoices:
            inv_type = inv.get("type", "")
            date_str = inv.get("date", "")
            if not date_str:
                continue
            try:
                d = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                continue

            # 滴滴行程单：打车地点中出现的每个城市都算
            if inv_type == "打车行程单":
                found_city = False
                for loc_key in ("start_location", "end_location"):
                    loc = inv.get(loc_key, "")
                    city = self._extract_city_from_location(loc)
                    if city and city != home_city and len(city) >= 2:
                        date_cities.setdefault(d, set()).add(city)
                        found_city = True
                if not found_city:
                    date_no_city.setdefault(d, []).append(inv)

            # 火车票/飞机票：出发城市和到达城市都算
            elif inv_type in ("火车票", "飞机票"):
                start = inv.get("start_location", "")
                end = inv.get("end_location", "")
                start_city = self._extract_city(start)
                end_city = self._extract_city(end)

                if end_city and end_city != home_city:
                    date_cities.setdefault(d, set()).add(end_city)
                if start_city and start_city != home_city:
                    date_cities.setdefault(d, set()).add(start_city)

        # ====== 第二步：城市继承（填补滴滴地点不含城市名的缺口） ======
        # 2a. 同一天内：有城市的滴滴记录传染给无城市的记录
        for d, invs in date_no_city.items():
            known = date_cities.get(d, set())
            if known:
                for city in known:
                    date_cities.setdefault(d, set()).add(city)

        # 2b. 相邻日期继承：如果某天的滴滴记录无城市，但前后1天有已知非家乡城市，则继承
        all_dates = sorted(set(list(date_cities.keys()) + list(date_no_city.keys())))
        for d in all_dates:
            if d in date_no_city and d not in date_cities:
                # 尝试从前1天继承
                prev_d = d - timedelta(days=1)
                if prev_d in date_cities:
                    for city in date_cities[prev_d]:
                        date_cities.setdefault(d, set()).add(city)
                    continue
                # 尝试从后1天继承
                next_d = d + timedelta(days=1)
                if next_d in date_cities:
                    for city in date_cities[next_d]:
                        date_cities.setdefault(d, set()).add(city)

        # ====== 第三步：每个城市独立计算出差天数 ======
        city_dates: Dict[str, set] = {}  # 翻转: {city: {date, ...}}
        for d, cities in date_cities.items():
            for city in cities:
                city_dates.setdefault(city, set()).add(d)

        # 每个城市独立计算出差天数
        travel_days: Dict[str, int] = {}
        for city, dates in city_dates.items():
            sorted_dates = sorted(dates)
            if not sorted_dates:
                continue

            # 按间隔≥3天分段（允许1-2天证据缺口被吸收）
            total = 0
            seg_start = sorted_dates[0]
            prev = sorted_dates[0]
            for d in sorted_dates[1:]:
                gap = (d - prev).days
                if gap >= 3:
                    total += (prev - seg_start).days + 1
                    seg_start = d
                prev = d
            total += (prev - seg_start).days + 1
            travel_days[city] = total

        # 酒店发票补充
        for inv in invoices:
            if inv.get("type") == "酒店":
                check_in = inv.get("check_in_date", "")
                check_out = inv.get("check_out_date", "")
                if check_in and check_out:
                    try:
                        d_in = datetime.strptime(check_in, "%Y-%m-%d").date()
                        d_out = datetime.strptime(check_out, "%Y-%m-%d").date()
                        nights = (d_out - d_in).days
                        if nights > 0:
                            hotel_city = self._infer_hotel_city(inv, invoices)
                            if hotel_city:
                                travel_days[hotel_city] = max(
                                    travel_days.get(hotel_city, 0), nights
                                )
                    except ValueError:
                        pass

        return travel_days

    # ------------------------------------------------------------------
    #  内部辅助方法
    # ------------------------------------------------------------------
    def _write_transport_row(self, row: int, inv: Dict):
        """写入一条城际交通记录（火车票 / 打车 / 飞机票）"""
        self.ws.cell(row=row, column=1, value=self._fmt_date(inv.get("date", "")))
        self.ws.cell(row=row, column=2, value=inv.get("start_location", ""))
        self.ws.cell(row=row, column=3, value=inv.get("end_location", ""))
        self.ws.cell(row=row, column=4, value=inv.get("work_content", ""))
        self.ws.cell(row=row, column=5, value=inv.get("type", ""))
        self.ws.cell(row=row, column=6, value=float(inv.get("amount", 0)))
        # 顺风车/拼车需要餐费代替
        remark = "餐费代替" if inv.get("need_substitute", False) else ""
        self.ws.cell(row=row, column=7, value=remark)

    def _write_hotel_row(self, row: int, inv: Dict):
        """写入一条酒店记录"""
        check_in = inv.get("check_in_date", "")
        check_out = inv.get("check_out_date", "")
        if check_in and check_out:
            time_range = f"{self._fmt_date(check_in)}-{self._fmt_date(check_out)}"
        else:
            time_range = self._fmt_date(inv.get("date", ""))

        self.ws.cell(row=row, column=1, value=time_range)
        self.ws.cell(row=row, column=2,
                      value=inv.get("hotel_name", inv.get("start_location", "")))
        nights = inv.get("nights", "")
        self.ws.cell(row=row, column=3, value=int(nights) if nights else "")
        daily_rate = inv.get("daily_rate", "")
        self.ws.cell(row=row, column=4,
                      value=float(daily_rate) if daily_rate else "")
        self.ws.cell(row=row, column=5, value="酒店")
        self.ws.cell(row=row, column=6, value=float(inv.get("amount", 0)))
        self.ws.cell(row=row, column=7, value="")

    def _write_other_row(self, row: int, inv: Dict):
        """写入一条其他费用记录"""
        self.ws.cell(row=row, column=1, value=self._fmt_date(inv.get("date", "")))
        self.ws.cell(row=row, column=2, value="")
        self.ws.cell(row=row, column=3, value=inv.get("type", ""))
        self.ws.cell(row=row, column=4, value=inv.get("work_content", ""))
        # 费用名称：快递 → 快递费，其他 → 原始类型
        fee_name = inv.get("type", "")
        if fee_name == "快递":
            fee_name = "快递费"
        self.ws.cell(row=row, column=5, value=fee_name)
        self.ws.cell(row=row, column=6, value=float(inv.get("amount", 0)))
        self.ws.cell(row=row, column=7, value="")

    # ---------- 补贴 ----------
    def _generate_subsidies(self, travel_days: Dict) -> List[Dict]:
        """生成补贴数据"""
        total_days = sum(travel_days.values()) if travel_days else 0
        subs = []
        if total_days > 0:
            subs.append({
                "name": "出差补贴",
                "amount": total_days * 50,
                "remark": "餐费代替",
            })
        subs.append({"name": "通讯补贴", "amount": 200, "remark": "餐费代替"})
        subs.append({"name": "餐补", "amount": 500, "remark": "餐费代替"})
        return subs

    def _build_travel_summary(self, travel_days: Dict) -> str:
        """构建出差天数摘要文字，如 '桐庐（2天），衢州（7天），浦江（2天），共11天，'"""
        if not travel_days:
            return ""
        parts = []
        total = 0
        for city, days in travel_days.items():
            parts.append(f"{city}（{days}天）")
            total += days
        return "，".join(parts) + f"，共{total}天，"

    # ---------- 日期 / 城市辅助 ----------
    @staticmethod
    def _fmt_date(date_str: str) -> str:
        """将 'YYYY-MM-DD' 转为 'YYYY.M.D' 格式（与模板一致）"""
        if not date_str:
            return ""
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d").date()
            return f"{d.year}.{d.month}.{d.day}"
        except ValueError:
            return date_str

    @staticmethod
    def _extract_city(location: str) -> str:
        """从站点名称中提取城市简称"""
        if not location:
            return ""
        # 去掉常见站名后缀
        for suffix in ("西站", "东站", "南站", "北站", "站", "东", "西", "南", "北"):
            if location.endswith(suffix) and len(location) > len(suffix):
                return location[: -len(suffix)]
        return location

    @staticmethod
    def _extract_city_from_location(location: str) -> str:
        """
        从打车地点字符串中提取城市名（无硬编码列表）
        支持格式：
        - "温州北站-网约车上车点" -> "温州"
        - "金华市车站路" -> "金华"
        - "桐庐县桐君街道" -> "桐庐"
        - "瓯海区|温州南站" -> "瓯海"
        """
        if not location:
            return ""
        import re
        # 匹配 "XX市" 格式（最可靠）
        m = re.search(r'([\u4e00-\u9fff]{2,4})市', location)
        if m:
            return m.group(1)
        # 匹配 "XX站" 格式（火车站名，去掉后缀提取城市）
        m = re.search(r'([\u4e00-\u9fff]{2,6})(?:西站|东站|南站|北站)', location)
        if m:
            return m.group(1)
        # 匹配 "XX区" 或 "XX县" 格式
        m = re.search(r'([\u4e00-\u9fff]{2,4})(?:区|县)', location)
        if m:
            return m.group(1)
        # 从括号中提取城市（如"全季酒店(温州车站大道店)" -> "温州"）
        m = re.search(r'[\(\uff08]([^\)\uff09]*?)[\)\uff09]', location)
        if m:
            inner = m.group(1)
            # 括号内容中查找“XX市/区/县”格式
            cm = re.search(r'([\u4e00-\u9fff]{2,4})(?:市|区|县)', inner)
            if cm:
                return cm.group(1)
            # 括号内容中查找“XX站”格式（排除“车站”“站台”等常见词）
            cm = re.search(r'([\u4e00-\u9fff]{2,6})(?:西站|东站|南站|北站|(?<!车)站)', inner)
            if cm:
                return cm.group(1)
            # 取括号内前2个中文字符（中国城市名通常为2字）
            cm = re.match(r'([\u4e00-\u9fff]{2})(?=[\u4e00-\u9fff])', inner)
            if cm:
                return cm.group(1)
        return ""

    def _infer_hotel_city(self, hotel: Dict, all_invoices: List[Dict]) -> str:
        """根据酒店名称和打车记录推断所在城市"""
        hotel_name = hotel.get("hotel_name", hotel.get("start_location", ""))
        check_in = hotel.get("check_in_date", "")
        if not hotel_name or not check_in:
            return ""

        try:
            check_in_date = datetime.strptime(check_in, "%Y-%m-%d").date()
        except ValueError:
            check_in_date = None

        # 在打车行程中找：目的地包含酒店关键字，且日期在入住期间
        for inv in all_invoices:
            if inv.get("type") == "打车" and inv.get("date"):
                end_loc = inv.get("end_location", "")
                if hotel_name and any(
                    kw in end_loc for kw in hotel_name.split("酒店")[0:1]
                ):
                    # 从同期火车票推断城市
                    inv_date = inv.get("date", "")
                    for t in all_invoices:
                        if t.get("type") == "火车票" and t.get("date") == inv_date:
                            start = t.get("start_location", "")
                            end = t.get("end_location", "")
                            if "杭州" in start:
                                return self._extract_city(end)
                            elif "杭州" in end:
                                return self._extract_city(start)

        # fallback: 找入住日期当天或前一天出发的火车票目的地
        if check_in_date:
            for inv in all_invoices:
                if inv.get("type") == "火车票" and inv.get("date"):
                    try:
                        t_date = datetime.strptime(inv["date"], "%Y-%m-%d").date()
                    except ValueError:
                        continue
                    if abs((t_date - check_in_date).days) <= 1:
                        start = inv.get("start_location", "")
                        end = inv.get("end_location", "")
                        if "杭州" in start:
                            return self._extract_city(end)

        return ""
