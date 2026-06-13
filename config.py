"""配置管理模块"""
import os
import io
from dotenv import load_dotenv

# 加载 .env 文件
load_dotenv()

# 通义千问 API 配置
DASHSCOPE_API_KEY = os.getenv("DASHSCOPE_API_KEY", "sk-70a3079eba7e462eb33f90ea77d9c51a")
DEFAULT_MODEL = "qwen-vl-max"

# 支持的费用类型
EXPENSE_TYPES = [
    "火车票",
    "打车",
    "飞机票",
    "酒店",
    "餐饮",
    "快递",
    "出差补贴",
    "通讯补贴",
    "餐补",
    "其他",
]

# 板块顺序
SECTION_ORDER = [
    "城际交通",
    "补贴",
    "住宿费",
    "其他费用",
]

# 费用类型到板块的映射
EXPENSE_SECTION_MAP = {
    "火车票": "城际交通",
    "打车": "城际交通",
    "打车行程单": "城际交通",
    "飞机票": "城际交通",
    "酒店": "住宿费",
    "餐饮": "其他费用",
    "出差补贴": "补贴",
    "通讯补贴": "补贴",
    "餐补": "补贴",
    "快递": "其他费用",
    "其他": "其他费用",
}

# 默认提示词文件路径
DEFAULT_PROMPT_PATH = os.path.join(
    os.path.dirname(__file__), "prompts", "default_prompt.txt"
)
INVOICE_PARSE_PROMPT_PATH = os.path.join(
    os.path.dirname(__file__), "prompts", "invoice_parse.txt"
)
INVOICE_PARSE_TEXT_PROMPT_PATH = os.path.join(
    os.path.dirname(__file__), "prompts", "invoice_parse_text.txt"
)
HOTEL_INFER_PROMPT_PATH = os.path.join(
    os.path.dirname(__file__), "prompts", "hotel_infer.txt"
)
WORK_MATCH_PROMPT_PATH = os.path.join(
    os.path.dirname(__file__), "prompts", "work_match.txt"
)

PROMPT_FILES = {
    "default_prompt": DEFAULT_PROMPT_PATH,
    "invoice_parse": INVOICE_PARSE_PROMPT_PATH,
    "invoice_parse_text": INVOICE_PARSE_TEXT_PROMPT_PATH,
    "hotel_infer": HOTEL_INFER_PROMPT_PATH,
    "work_match": WORK_MATCH_PROMPT_PATH,
}

PROMPT_NAMES = {
    "default_prompt": "报销规则总提示词",
    "invoice_parse": "发票识别提示词（视觉AI）",
    "invoice_parse_text": "发票识别提示词（文本解析）",
    "hotel_infer": "酒店日期推算提示词",
    "work_match": "工作内容匹配提示词",
}


def load_prompt(prompt_key: str) -> str:
    """加载指定提示词文件"""
    path = PROMPT_FILES.get(prompt_key)
    if not path:
        return ""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return ""


def get_api_key(session_api_key: str = "") -> str:
    """获取API Key，优先使用页面输入的值"""
    if session_api_key:
        return session_api_key
    return DASHSCOPE_API_KEY


# 内置空白报销模板路径
BUILTIN_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "templates", "blank_template.xlsx"
)


def get_builtin_template_bytes() -> bytes:
    """加载内置空白报销模板（动态生成当月标题）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # 优先读取静态模板文件
    try:
        with open(BUILTIN_TEMPLATE_PATH, "rb") as f:
            static_bytes = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(static_bytes))
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '报销单'

    # 动态更新标题月份
    month = datetime.now().month
    ws['A1'] = f'{month}月 浙江地区行程报销单'

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
